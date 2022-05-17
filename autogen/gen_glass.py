import openpyxl
import os, sys
import json

source_fn = '../data/LaCroix Dynamic Material Selection Data Tool vJanuary 2015.xlsm'
output_fn = '../obj_correct/data/glass.json'

if not os.path.exists(source_fn):
    print('''This script generates glass data from a publicly avaiable spreadsheet, which
you should place in the ../data directory.  This sheet can be found at:
https://web.archive.org/web/20151011033820/http://www.lacroixoptical.com/sites/default/files/content/LaCroix%20Dynamic%20Material%20Selection%20Data%20Tool%20vJanuary%202015.xlsm

(Also linked from: https://en.wikipedia.org/wiki/Sellmeier_equation)
''')

data = openpyxl.load_workbook(source_fn)['AllData']
headings = {data.cell(row=2, column=n).value:n
            for n in range(1, 100) if data.cell(row=2, column=n).value}

glass_data = {
    data.cell(row=i, column=headings['Glass Type']).value: {
        'B':[data.cell(row=i, column=headings[f'B{j}']).value for j in range(1, 4)],
        'C':[data.cell(row=i, column=headings[f'C{j}']).value for j in range(1, 4)],
    } for i in range(3, 1000) if data.cell(row=i, column=headings['Glass Type']).value
}

# Manually add some useful ones!

# https://opg.optica.org/ao/viewmedia.cfm?uri=ao-46-18-3811&seq=0&html=true
# used values for 20C
glass_data['water'] = {
    'B':[5.684027565E-1, 1.726177391E-1, 2.086189578E-2, 1.130748688E-1],
    'C':[5.101829712E-3, 1.821153936E-2, 2.620722293E-2, 1.069792721E1]
}

# https://en.wikipedia.org/wiki/Sellmeier_equation
glass_data['FS'] = {
    'B':[0.696166300, 0.407942600, 0.897479400],
    'C':[4.67914826E-3, 1.35120631E-2, 97.9340025]
}

# https://www.corning.com/media/worldwide/csm/documents/HPFS_Product_Brochure_All_Grades_2015_07_21.pdf
glass_data['Fused Silica (Corning 7980)'] = {
    'B':[0.68374049400, 0.42032361300, 0.58502748000],
    'C':[0.00460352869, 0.01339688560, 64.49327320000]
}

glass_data['7980'] = glass_data['Fused Silica (Corning 7980)']

with open(output_fn, 'w') as f:
    json.dump(glass_data, f)
