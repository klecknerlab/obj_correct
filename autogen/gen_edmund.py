import openpyxl
import os, sys
import json

source_fn = '../data/VIS-EXT-Coated-Plano-Convex-PCX-Lenses.xlsx'
output_fn = '../obj_correct/data/edmund_plano_convex.json'

if not os.path.exists(source_fn):
    print('''This script generates glass data from a publicly avaiable spreadsheet, which
you should place in the ../data directory.  This sheet can be found at:
https://www.edmundoptics.com/f/vis-ext-coated-plano-convex-pcx-lenses/14893/
(click on "Expoer Specifications" button)
''')

data = openpyxl.load_workbook(source_fn)['3679']
headings = {data.cell(row=1, column=n).value:n
            for n in range(1, 100) if data.cell(row=2, column=n).value}

k = headings['Stock Number']

def strip_unc(s, k=''):
    if 'mm' in k:
        for c in ('+', '±', '@', '-'):
            if c in s:
                s = s[:s.index(c)]
        return float(s)
    else:
        return s.strip()

fields = {
    'description':'Title',
    'diameter': 'Dia. (mm)',
    'glass': 'Substrate',
    'f': 'EFL (mm)',
    'CA': 'CA (mm)',
    't': 'CT (mm)',
    'R1': 'Radius R1 (mm)'
}

lens_data = {
    data.cell(row=i, column=k).value: {
        k:strip_unc(data.cell(row=i, column=headings[v]).value, v) for k, v in fields.items()
    } for i in range(3, 1000) if data.cell(row=i, column=k).value
}

with open(output_fn, 'w') as f:
    json.dump(lens_data, f)
